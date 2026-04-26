"""URL API routes — crawled URL queries, inlinks/outlinks, CSV/sitemap export."""

from __future__ import annotations

import csv
import io
import re
import uuid
from datetime import datetime, timezone
from urllib.parse import urlparse
from xml.etree import ElementTree as ET
from xml.sax.saxutils import escape as xml_escape

import aiohttp
import structlog
from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import StreamingResponse

from app.api.deps import DbSession
from app.repositories.url_repo import UrlRepository
from app.schemas.pagination import CursorPage
from app.schemas.url import CrawledUrlDetail, CrawledUrlResponse, ExternalLinkResponse

logger = structlog.get_logger(__name__)

router = APIRouter(tags=["urls"])


@router.get(
    "/crawls/{crawl_id}/urls",
    response_model=CursorPage[CrawledUrlResponse],
)
async def list_crawled_urls(
    crawl_id: uuid.UUID,
    db: DbSession,
    cursor: str | None = Query(None),
    limit: int = Query(50, ge=1, le=500),
    status_code: int | None = Query(None, description="Filter by HTTP status code"),
    content_type: str | None = Query(None, description="Filter by content type (partial match)"),
    is_indexable: bool | None = Query(None, description="Filter by indexability"),
    search: str | None = Query(None, description="Search URL text (ILIKE)"),
    status_code_min: int | None = Query(None, description="Min status code (inclusive)"),
    status_code_max: int | None = Query(None, description="Max status code (inclusive)"),
    has_issue: str | None = Query(None, description="Filter by issue type"),
) -> CursorPage[CrawledUrlResponse]:
    repo = UrlRepository(db)
    result = await repo.list_by_crawl(
        crawl_id=crawl_id,
        cursor=cursor,
        limit=limit,
        status_code=status_code,
        content_type=content_type,
        is_indexable=is_indexable,
        search=search,
        status_code_min=status_code_min,
        status_code_max=status_code_max,
        has_issue=has_issue,
    )
    return CursorPage(
        items=[CrawledUrlResponse.model_validate(u) for u in result["items"]],
        next_cursor=result["next_cursor"],
    )


@router.get(
    "/crawls/{crawl_id}/urls/{url_id}",
    response_model=CrawledUrlDetail,
)
async def get_crawled_url(
    crawl_id: uuid.UUID,
    url_id: uuid.UUID,
    db: DbSession,
) -> CrawledUrlDetail:
    repo = UrlRepository(db)
    url = await repo.get_by_id(url_id=url_id, crawl_id=crawl_id)
    if url is None:
        raise HTTPException(status_code=404, detail="URL not found")
    return CrawledUrlDetail.model_validate(url)


@router.get("/crawls/{crawl_id}/urls/{url_id}/inlinks")
async def get_url_inlinks(
    crawl_id: uuid.UUID,
    url_id: uuid.UUID,
    db: DbSession,
    limit: int = Query(100, ge=1, le=500),
) -> list[dict]:
    repo = UrlRepository(db)
    return await repo.get_inlinks(crawl_id=crawl_id, url_id=url_id, limit=limit)


@router.get("/crawls/{crawl_id}/urls/{url_id}/outlinks")
async def get_url_outlinks(
    crawl_id: uuid.UUID,
    url_id: uuid.UUID,
    db: DbSession,
    limit: int = Query(100, ge=1, le=500),
) -> list[dict]:
    repo = UrlRepository(db)
    return await repo.get_outlinks(crawl_id=crawl_id, url_id=url_id, limit=limit)


@router.get(
    "/crawls/{crawl_id}/external-links",
    response_model=CursorPage[ExternalLinkResponse],
)
async def list_external_links(
    crawl_id: uuid.UUID,
    db: DbSession,
    cursor: str | None = Query(None),
    limit: int = Query(50, ge=1, le=500),
    search: str | None = Query(None),
    nofollow: bool | None = Query(None),
) -> CursorPage[ExternalLinkResponse]:
    repo = UrlRepository(db)
    int_cursor = None
    if cursor:
        try:
            int_cursor = int(cursor)
        except (ValueError, TypeError):
            return CursorPage(items=[], next_cursor=None)
    result = await repo.list_external_links(
        crawl_id=crawl_id,
        cursor=int_cursor,
        limit=limit,
        search=search,
        nofollow=nofollow,
    )
    return CursorPage(
        items=[ExternalLinkResponse.model_validate(u) for u in result["items"]],
        next_cursor=result["next_cursor"],
    )


@router.get("/crawls/{crawl_id}/sitemap.xml")
async def generate_sitemap(
    crawl_id: uuid.UUID,
    db: DbSession,
    include_non_indexable: bool = Query(False),
    include_non_200: bool = Query(False),
) -> StreamingResponse:
    repo = UrlRepository(db)

    async def _generate_xml():
        yield '<?xml version="1.0" encoding="UTF-8"?>\n'
        yield '<urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9">\n'

        async for batch in repo.stream_for_sitemap(
            crawl_id=crawl_id,
            include_non_indexable=include_non_indexable,
            include_non_200=include_non_200,
        ):
            for url_obj in batch:
                loc = xml_escape(url_obj.url)
                lastmod = url_obj.crawled_at.strftime("%Y-%m-%d") if url_obj.crawled_at else ""
                yield "  <url>\n"
                yield f"    <loc>{loc}</loc>\n"
                if lastmod:
                    yield f"    <lastmod>{lastmod}</lastmod>\n"
                freq = _change_freq(url_obj.crawl_depth)
                priority = _priority(url_obj.crawl_depth)
                yield f"    <changefreq>{freq}</changefreq>\n"
                yield f"    <priority>{priority}</priority>\n"
                yield "  </url>\n"

        yield "</urlset>\n"

    return StreamingResponse(
        _generate_xml(),
        media_type="application/xml",
        headers={"Content-Disposition": f"attachment; filename=sitemap_{crawl_id}.xml"},
    )


def _change_freq(depth: int) -> str:
    if depth == 0:
        return "daily"
    if depth <= 2:
        return "weekly"
    return "monthly"


def _priority(depth: int) -> str:
    if depth == 0:
        return "1.0"
    if depth == 1:
        return "0.8"
    if depth == 2:
        return "0.6"
    if depth <= 4:
        return "0.4"
    return "0.2"


EXPORT_COLUMNS = [
    "url",
    "status_code",
    "content_type",
    "title",
    "title_length",
    "title_pixel_width",
    "meta_description",
    "meta_desc_length",
    "h1",
    "h2",
    "canonical_url",
    "robots_meta",
    "is_indexable",
    "indexability_reason",
    "word_count",
    "crawl_depth",
    "response_time_ms",
    "redirect_url",
    "link_score",
    "text_ratio",
    "readability_score",
    "avg_words_per_sentence",
]


@router.get("/crawls/{crawl_id}/export")
async def export_crawl_csv(
    crawl_id: uuid.UUID,
    db: DbSession,
) -> StreamingResponse:
    repo = UrlRepository(db)

    async def _generate_csv():
        # Header row
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(EXPORT_COLUMNS)
        yield buf.getvalue()

        # Stream data in batches
        async for batch in repo.stream_for_export(crawl_id=crawl_id, batch_size=500):
            buf = io.StringIO()
            writer = csv.writer(buf)
            for u in batch:
                row = []
                for col in EXPORT_COLUMNS:
                    val = getattr(u, col, None)
                    if isinstance(val, list):
                        val = " | ".join(str(v) for v in val)
                    row.append(val if val is not None else "")
                writer.writerow(row)
            yield buf.getvalue()

    return StreamingResponse(
        _generate_csv(),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=crawl_{crawl_id}.csv"},
    )


@router.get("/crawls/{crawl_id}/export-xlsx")
async def export_crawl_xlsx(
    crawl_id: uuid.UUID,
    db: DbSession,
) -> StreamingResponse:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    repo = UrlRepository(db)

    wb = Workbook(write_only=True)
    ws = wb.create_sheet("Crawl Data")

    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="6CC04A", end_color="6CC04A", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    from openpyxl.cell import WriteOnlyCell
    header_cells = []
    for col_name in EXPORT_COLUMNS:
        cell = WriteOnlyCell(ws, value=col_name.replace("_", " ").title())
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        header_cells.append(cell)
    ws.append(header_cells)

    # Set column widths based on header lengths
    for col_idx, col_name in enumerate(EXPORT_COLUMNS, 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = min(len(col_name) + 4, 40)

    row_num = 2
    async for batch in repo.stream_for_export(crawl_id=crawl_id, batch_size=500):
        for u in batch:
            row_data = []
            for col in EXPORT_COLUMNS:
                val = getattr(u, col, None)
                if isinstance(val, list):
                    val = " | ".join(str(v) for v in val)
                row_data.append(val if val is not None else "")
            ws.append(row_data)
            row_num += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=crawl_{crawl_id}.xlsx"},
    )


REQUIRED_FIELDS: dict[str, list[str]] = {
    "Article": ["headline", "author", "datePublished"],
    "NewsArticle": ["headline", "author", "datePublished"],
    "BlogPosting": ["headline", "author", "datePublished"],
    "Product": ["name"],
    "LocalBusiness": ["name", "address"],
    "Organization": ["name"],
    "Person": ["name"],
    "WebSite": ["name", "url"],
    "WebPage": ["name"],
    "BreadcrumbList": ["itemListElement"],
    "FAQPage": ["mainEntity"],
    "HowTo": ["name", "step"],
    "Recipe": ["name"],
    "Event": ["name", "startDate", "location"],
    "VideoObject": ["name", "uploadDate"],
    "ImageObject": ["contentUrl"],
    "Review": ["itemReviewed"],
    "AggregateRating": ["ratingValue", "reviewCount"],
}


def _validate_json_ld_block(block: dict) -> list[dict]:
    issues: list[dict] = []
    schema_type = block.get("@type")
    if not schema_type:
        issues.append({"level": "error", "message": "Missing @type property"})
        return issues
    if not block.get("@context"):
        issues.append(
            {"level": "warning", "message": "Missing @context (should be https://schema.org)"}
        )

    types = schema_type if isinstance(schema_type, list) else [schema_type]
    for t in types:
        required = REQUIRED_FIELDS.get(t, [])
        for field in required:
            if field not in block or block[field] is None or block[field] == "":
                issues.append(
                    {"level": "warning", "message": f"{t}: missing recommended field '{field}'"}
                )
    return issues


@router.get("/crawls/{crawl_id}/structured-data")
async def list_structured_data(
    crawl_id: uuid.UUID,
    db: DbSession,
    cursor: str | None = Query(None),
    limit: int = Query(50, ge=1, le=200),
) -> dict:
    repo = UrlRepository(db)
    result = await repo.list_with_structured_data(crawl_id=crawl_id, cursor=cursor, limit=limit)

    items = []
    for url_obj in result["items"]:
        json_ld = (url_obj.seo_data or {}).get("json_ld", [])
        blocks_with_validation = []
        for block in json_ld if isinstance(json_ld, list) else []:
            if not isinstance(block, dict):
                continue
            validation = _validate_json_ld_block(block)
            schema_type = block.get("@type", "Unknown")
            if isinstance(schema_type, list):
                schema_type = ", ".join(schema_type)
            blocks_with_validation.append(
                {
                    "type": schema_type,
                    "data": block,
                    "issues": validation,
                    "is_valid": all(v["level"] != "error" for v in validation),
                }
            )
        items.append(
            {
                "url_id": str(url_obj.id),
                "url": url_obj.url,
                "blocks": blocks_with_validation,
                "block_count": len(blocks_with_validation),
                "has_errors": any(not b["is_valid"] for b in blocks_with_validation),
            }
        )

    return {
        "items": items,
        "next_cursor": result["next_cursor"],
    }


@router.get("/crawls/{crawl_id}/custom-extractions")
async def list_custom_extractions(
    crawl_id: uuid.UUID,
    db: DbSession,
    cursor: str | None = Query(None),
    limit: int = Query(50, ge=1, le=200),
) -> dict:
    repo = UrlRepository(db)
    result = await repo.list_with_custom_extractions(
        crawl_id=crawl_id,
        cursor=cursor,
        limit=limit,
    )
    items = []
    for url_obj in result["items"]:
        seo_data = url_obj.seo_data or {}
        extractions = seo_data.get("custom_extractions", {})
        if extractions:
            items.append(
                {
                    "url_id": str(url_obj.id),
                    "url": url_obj.url,
                    "extractions": extractions,
                }
            )
    return {
        "items": items,
        "next_cursor": result["next_cursor"],
    }


@router.get("/crawls/{crawl_id}/pagination")
async def list_pagination_urls(
    crawl_id: uuid.UUID,
    db: DbSession,
    filter: str | None = Query(None, description="Filter type: contains, first_page, paginated_2_plus, url_not_in_anchor, non_200, unlinked, non_indexable, multiple, loop, sequence_error"),
    cursor: str | None = Query(None),
    limit: int = Query(50, ge=1, le=200),
) -> dict:
    """List URLs with pagination rel=next/prev attributes."""
    repo = UrlRepository(db)
    result = await repo.list_pagination_urls(
        crawl_id=crawl_id,
        filter_type=filter,
        cursor=cursor,
        limit=limit,
    )
    items = []
    for url_obj in result["items"]:
        seo_data = url_obj.seo_data or {}
        pag = seo_data.get("pagination", {})
        items.append(
            {
                "url_id": str(url_obj.id),
                "url": url_obj.url,
                "status_code": url_obj.status_code,
                "rel_next": pag.get("rel_next"),
                "rel_prev": pag.get("rel_prev"),
                "is_indexable": url_obj.is_indexable,
                "indexability_reason": url_obj.indexability_reason,
            }
        )
    return {
        "items": items,
        "next_cursor": result["next_cursor"],
    }


@router.get("/crawls/{crawl_id}/links/analysis")
async def get_links_analysis(
    crawl_id: uuid.UUID,
    db: DbSession,
    limit: int = Query(50, ge=1, le=500),
) -> dict:
    """Internal links analysis: inlink counts, orphan pages, depth distribution, anchor text stats."""
    repo = UrlRepository(db)
    return {
        "top_pages_by_inlinks": await repo.get_inlink_counts(crawl_id, limit=limit),
        "orphan_pages": await repo.get_orphan_pages(crawl_id, limit=limit),
        "depth_distribution": await repo.get_depth_distribution(crawl_id),
        "anchor_text_stats": await repo.get_anchor_text_stats(crawl_id, limit=50),
    }


@router.get("/crawls/{crawl_id}/duplicates")
async def get_duplicates(
    crawl_id: uuid.UUID,
    db: DbSession,
) -> dict:
    """Get exact and near-duplicate URL groups."""
    repo = UrlRepository(db)
    return {
        "exact_duplicates": await repo.get_exact_duplicate_groups(crawl_id),
        "near_duplicates": await repo.get_near_duplicate_groups(crawl_id),
    }



@router.get("/crawls/{crawl_id}/content-analysis")
async def get_content_analysis(
    crawl_id: uuid.UUID,
    db: DbSession,
    limit: int = Query(100, ge=1, le=1000),
) -> list[dict]:
    """Get content analysis metrics: readability, text ratio, word count."""
    repo = UrlRepository(db)
    return await repo.get_content_analysis(crawl_id, limit=limit)


@router.get("/crawls/{crawl_id}/link-scores")
async def get_link_scores(
    crawl_id: uuid.UUID,
    db: DbSession,
    limit: int = Query(100, ge=1, le=1000),
) -> list[dict]:
    """Get URLs ranked by Link Score (internal PageRank)."""
    repo = UrlRepository(db)
    return await repo.get_link_scores(crawl_id, limit=limit)


@router.get("/crawls/{crawl_id}/redirects")
async def get_redirect_chains(
    crawl_id: uuid.UUID,
    db: DbSession,
    limit: int = Query(200, ge=1, le=1000),
) -> list[dict]:
    """Get pages with redirects and their full redirect chains."""
    repo = UrlRepository(db)
    return await repo.get_redirect_chains(crawl_id, limit=limit)


@router.get("/crawls/{crawl_id}/health")
async def get_health_score(
    crawl_id: uuid.UUID,
    db: DbSession,
) -> dict:
    """Get SEO health score (0-100) with component breakdown."""
    repo = UrlRepository(db)
    return await repo.get_health_score(crawl_id)


@router.get("/crawls/{crawl_id}/performance")
async def get_performance_stats(
    crawl_id: uuid.UUID,
    db: DbSession,
    limit: int = Query(50, ge=1, le=500),
) -> dict:
    """Get response time statistics and slowest pages."""
    repo = UrlRepository(db)
    return await repo.get_performance_stats(crawl_id, limit=limit)


@router.get("/crawls/{crawl_id}/cookies")
async def get_cookies_audit(
    crawl_id: uuid.UUID,
    db: DbSession,
    limit: int = Query(200, ge=1, le=1000),
) -> list[dict]:
    """Get cookie audit: pages with Set-Cookie headers and security flags."""
    repo = UrlRepository(db)
    return await repo.get_cookies_audit(crawl_id, limit=limit)


@router.get("/crawls/{crawl_id}/security")
async def get_security_overview(
    crawl_id: uuid.UUID,
    db: DbSession,
) -> dict:
    """Get security overview: HTTPS adoption, security header coverage."""
    repo = UrlRepository(db)
    return await repo.get_security_overview(crawl_id)


@router.get("/crawls/{crawl_id}/hreflang")
async def get_hreflang_data(
    crawl_id: uuid.UUID,
    db: DbSession,
    limit: int = Query(200, ge=1, le=1000),
) -> list[dict]:
    """Get pages with hreflang tags."""
    repo = UrlRepository(db)
    return await repo.get_hreflang_data(crawl_id, limit=limit)


@router.get("/crawls/{crawl_id}/heading-hierarchy")
async def get_heading_hierarchy(
    crawl_id: uuid.UUID,
    db: DbSession,
    limit: int = Query(200, ge=1, le=1000),
) -> list[dict]:
    """Get heading hierarchy analysis: sequence, skip-level detection."""
    repo = UrlRepository(db)
    return await repo.get_heading_hierarchy(crawl_id, limit=limit)


@router.get("/crawls/{crawl_id}/overview-stats")
async def get_overview_stats(
    crawl_id: uuid.UUID,
    db: DbSession,
) -> dict:
    """Get comprehensive overview stats: status code dist, content type dist, indexability."""
    repo = UrlRepository(db)
    return await repo.get_overview_stats(crawl_id)


@router.get("/crawls/{crawl_id}/images-audit")
async def get_images_audit(
    crawl_id: uuid.UUID,
    db: DbSession,
    limit: int = Query(500, ge=1, le=2000),
) -> dict:
    """Get images audit: alt text coverage, missing dimensions, per-page breakdown."""
    repo = UrlRepository(db)
    return await repo.get_images_audit(crawl_id, limit=limit)


@router.get("/crawls/{crawl_id}/site-structure")
async def get_site_structure(
    crawl_id: uuid.UUID,
    db: DbSession,
    max_nodes: int = Query(500, ge=1, le=5000),
) -> dict:
    """Build site structure tree from crawled URLs."""
    repo = UrlRepository(db)
    urls_data = await repo.get_site_structure(crawl_id, limit=max_nodes)

    # Build tree from URL paths
    tree: dict = {"name": "/", "path": "/", "children": {}, "pages": 0, "status_codes": {}, "depth": 0}

    for u in urls_data:
        url_str = u["url"]
        try:
            parsed = urlparse(url_str)
            path = parsed.path.rstrip("/") or "/"
            parts = [p for p in path.split("/") if p]
        except Exception:
            continue

        node = tree
        node["pages"] += 1
        sc = str(u["status_code"] or "none")
        node["status_codes"][sc] = node["status_codes"].get(sc, 0) + 1

        for part in parts:
            if part not in node["children"]:
                current_path = "/" + "/".join(parts[:parts.index(part) + 1])
                node["children"][part] = {
                    "name": part,
                    "path": current_path,
                    "children": {},
                    "pages": 0,
                    "status_codes": {},
                    "depth": parts.index(part) + 1,
                }
            node = node["children"][part]
            node["pages"] += 1
            node["status_codes"][sc] = node["status_codes"].get(sc, 0) + 1

    # Convert children dicts to sorted lists
    def _flatten(n: dict) -> dict:
        children = sorted(n["children"].values(), key=lambda c: -c["pages"])
        return {
            "name": n["name"],
            "path": n["path"],
            "pages": n["pages"],
            "status_codes": n["status_codes"],
            "depth": n["depth"],
            "children": [_flatten(c) for c in children],
        }

    # Compute directory stats
    depth_counts: dict[int, int] = {}
    for u in urls_data:
        d = u.get("crawl_depth", 0)
        depth_counts[d] = depth_counts.get(d, 0) + 1

    return {
        "tree": _flatten(tree),
        "total_urls": len(urls_data),
        "depth_distribution": [{"depth": d, "count": c} for d, c in sorted(depth_counts.items())],
        "max_depth": max((u.get("crawl_depth", 0) for u in urls_data), default=0),
    }


async def _get_crawl_domain(db: DbSession, crawl_id: uuid.UUID) -> tuple[str, str]:
    """Get scheme and domain from crawl's start_url. Returns (scheme, domain)."""
    from app.models.crawl import Crawl

    result = await db.get(Crawl, crawl_id)
    if result is None:
        raise HTTPException(status_code=404, detail="Crawl not found")
    start_url = (result.config or {}).get("start_url", "")
    if not start_url:
        raise HTTPException(status_code=400, detail="Crawl has no start_url")
    parsed = urlparse(start_url)
    return parsed.scheme or "https", parsed.netloc


@router.get("/crawls/{crawl_id}/robots-txt")
async def get_robots_txt(
    crawl_id: uuid.UUID,
    db: DbSession,
) -> dict:
    """Fetch and parse robots.txt for the crawled domain."""
    scheme, domain = await _get_crawl_domain(db, crawl_id)
    robots_url = f"{scheme}://{domain}/robots.txt"

    raw_content = ""
    status_code = 0
    try:
        async with aiohttp.ClientSession() as session:
            async with session.get(robots_url, timeout=aiohttp.ClientTimeout(total=10)) as resp:
                status_code = resp.status
                if resp.status == 200:
                    raw_content = await resp.text(errors="replace")
    except Exception as e:
        logger.debug("robots_txt_fetch_error", url=robots_url, error=str(e))

    # Parse directives
    directives: list[dict] = []
    sitemaps: list[str] = []
    current_agent = "*"

    if raw_content:
        for line in raw_content.splitlines():
            stripped = line.strip()
            if not stripped or stripped.startswith("#"):
                continue
            match = re.match(r"^([\w-]+)\s*:\s*(.+)$", stripped, re.IGNORECASE)
            if not match:
                continue
            key = match.group(1).lower()
            value = match.group(2).strip()
            if key == "user-agent":
                current_agent = value
            elif key == "sitemap":
                sitemaps.append(value)
            else:
                directives.append({
                    "user_agent": current_agent,
                    "directive": key,
                    "value": value,
                })

    return {
        "url": robots_url,
        "status_code": status_code,
        "raw_content": raw_content,
        "directives": directives,
        "sitemaps": sitemaps,
        "domain": domain,
    }


@router.get("/crawls/{crawl_id}/sitemap-analysis")
async def get_sitemap_analysis(
    crawl_id: uuid.UUID,
    db: DbSession,
) -> dict:
    """Discover and analyze sitemaps, compare with crawled URLs."""
    scheme, domain = await _get_crawl_domain(db, crawl_id)

    # Step 1: Discover sitemap URLs from robots.txt
    sitemap_urls: list[str] = []
    try:
        async with aiohttp.ClientSession() as session:
            async with session.get(
                f"{scheme}://{domain}/robots.txt",
                timeout=aiohttp.ClientTimeout(total=10),
            ) as resp:
                if resp.status == 200:
                    text = await resp.text(errors="replace")
                    for line in text.splitlines():
                        match = re.match(r"^sitemap\s*:\s*(.+)$", line.strip(), re.IGNORECASE)
                        if match:
                            sitemap_urls.append(match.group(1).strip())
    except Exception:
        pass

    # Fallback: try common sitemap locations
    if not sitemap_urls:
        sitemap_urls = [f"{scheme}://{domain}/sitemap.xml"]

    # Step 2: Fetch and parse sitemaps (with sitemap index support)
    sitemap_entries: list[dict] = []
    parsed_sitemaps: list[dict] = []

    async def _fetch_sitemap(url: str, depth: int = 0) -> None:
        if depth > 2:  # Prevent infinite recursion
            return
        try:
            async with aiohttp.ClientSession() as session:
                async with session.get(
                    url, timeout=aiohttp.ClientTimeout(total=15)
                ) as resp:
                    if resp.status != 200:
                        parsed_sitemaps.append({
                            "url": url,
                            "status_code": resp.status,
                            "type": "error",
                            "url_count": 0,
                        })
                        return
                    content = await resp.text(errors="replace")
        except Exception as e:
            parsed_sitemaps.append({
                "url": url,
                "status_code": 0,
                "type": "error",
                "url_count": 0,
                "error": str(e),
            })
            return

        try:
            root = ET.fromstring(content)
        except ET.ParseError:
            parsed_sitemaps.append({
                "url": url,
                "status_code": 200,
                "type": "parse_error",
                "url_count": 0,
            })
            return

        ns = ""
        tag = root.tag
        if tag.startswith("{"):
            ns = tag[: tag.index("}") + 1]

        # Sitemap index
        if root.tag == f"{ns}sitemapindex":
            child_urls = []
            for sitemap_el in root.findall(f"{ns}sitemap"):
                loc_el = sitemap_el.find(f"{ns}loc")
                if loc_el is not None and loc_el.text:
                    child_urls.append(loc_el.text.strip())
            parsed_sitemaps.append({
                "url": url,
                "status_code": 200,
                "type": "index",
                "url_count": len(child_urls),
                "child_sitemaps": child_urls,
            })
            for child_url in child_urls[:10]:  # Limit to 10 child sitemaps
                await _fetch_sitemap(child_url, depth + 1)
        else:
            # Regular urlset
            urls_found = []
            for url_el in root.findall(f"{ns}url"):
                loc_el = url_el.find(f"{ns}loc")
                lastmod_el = url_el.find(f"{ns}lastmod")
                if loc_el is not None and loc_el.text:
                    entry = {
                        "url": loc_el.text.strip(),
                        "lastmod": lastmod_el.text.strip() if lastmod_el is not None and lastmod_el.text else None,
                    }
                    urls_found.append(entry)
                    sitemap_entries.append(entry)
            parsed_sitemaps.append({
                "url": url,
                "status_code": 200,
                "type": "urlset",
                "url_count": len(urls_found),
            })

    for sm_url in sitemap_urls:
        await _fetch_sitemap(sm_url)

    # Step 3: Compare with crawled URLs
    repo = UrlRepository(db)
    crawled_urls_set: set[str] = set()
    async for batch in repo.stream_for_sitemap(
        crawl_id=crawl_id,
        include_non_indexable=True,
        include_non_200=True,
    ):
        for u in batch:
            crawled_urls_set.add(u.url.rstrip("/"))

    sitemap_urls_set = {e["url"].rstrip("/") for e in sitemap_entries}

    in_sitemap_not_crawled = sorted(sitemap_urls_set - crawled_urls_set)
    in_crawl_not_sitemap = sorted(crawled_urls_set - sitemap_urls_set)
    in_both = sorted(sitemap_urls_set & crawled_urls_set)

    return {
        "domain": domain,
        "sitemaps": parsed_sitemaps,
        "total_sitemap_urls": len(sitemap_entries),
        "total_crawled_urls": len(crawled_urls_set),
        "coverage": {
            "in_both": len(in_both),
            "in_sitemap_not_crawled": len(in_sitemap_not_crawled),
            "in_crawl_not_sitemap": len(in_crawl_not_sitemap),
        },
        "urls_in_sitemap_not_crawled": in_sitemap_not_crawled[:200],
        "urls_in_crawl_not_sitemap": in_crawl_not_sitemap[:200],
    }
